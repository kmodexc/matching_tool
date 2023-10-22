import pandas as pd
import openpyxl
import numpy as np
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import maximum_flow
import networkx as nx
import matplotlib.pyplot as plt

DAYS = ['mon', 'tue', 'wed', 'thr', 'fri', 'sat', 'sun']
SHIFTS = ['9', '14', '18']
SHIFT_TYPES = ["cut", "glu", "san", "che", "vab", "lam", "cor"]


def read_sheet(file="Shiftsplan_lux025.xlsx", sheet="ShiftList_KW16"):
    dataframe = openpyxl.load_workbook(file)
    dataframe1 = dataframe[sheet]
    name = None
    n_shifts = None

    data = {'name': [], 'n_shifts': [], 'beer': []}

    for d in DAYS:
        for s in SHIFTS:
            key = d + "_" + s
            data[key] = []

    for row in range(13, dataframe1.max_row):
        for i, col in enumerate(dataframe1.iter_cols(2, 24)):
            if i == 0:
                name = col[row].value
                if name != None:
                    data['name'].append(name)
                #     data['shiftlead'].append(col[row].font is not None and col[row].font.color is not None \
                #         and str(col[row].font.color.rgb) != "Values must be of type <class 'str'>" \
                #         and col[row].font.color.rgb != "FF000000")
            elif i == 1:
                if name != None:
                    n_shifts = col[row].value
                    data['beer'].append(1 if n_shifts is None else 0)
                    if n_shifts is None:
                        n_shifts = 0
                    data['n_shifts'].append(int(n_shifts))
            elif name != None:
                key = DAYS[(i - 2) // len(SHIFTS)] + "_" + SHIFTS[(i - 2) % len(SHIFTS)]
                if col[row].value != None:
                    data[key].append(1)
                else:
                    data[key].append(0)

    for st in SHIFT_TYPES:
        data["sl_" + st] = len(data["name"]) * [False]

    df = pd.DataFrame(data)

    for row in range(1, dataframe["ShiftLeads"].max_row):
        for i, col in enumerate(dataframe["ShiftLeads"].iter_cols(1, 8)):
            if i == 0:
                name = col[row].value
            elif name != None and name != "" and col[row].value != None and len(str(col[row].value)) > 0:
                df.loc[df["name"] == name, "sl_" + SHIFT_TYPES[i - 2]] = True

    shift_type_dict = {}
    for row in range(5, 12):
        for i, col in enumerate(dataframe1.iter_cols(4, 24)):
            key = DAYS[(i - 3) // len(SHIFTS)] + "_" + SHIFTS[(i - 3) % len(SHIFTS)]
            shift_type_id = SHIFT_TYPES[row - 5]
            if key not in shift_type_dict:
                shift_type_dict[key] = []
            if col[row].value is not None:
                shift_type_dict[key].append(shift_type_id)

    return df, shift_type_dict


def shiftplan_to_adj_mat(df):
    n_nodes = 1 + df.shape[0] + (len(SHIFTS) * len(DAYS)) + 1
    graph_matrix = np.zeros((n_nodes, n_nodes), dtype="int64")
    graph_matrix[0, 1:(1 + df.shape[0])] = df["n_shifts"].to_numpy()
    graph_matrix[(1 + df.shape[0]):-1, -1] = 3
    for i, d in enumerate(DAYS):
        for j, s in enumerate(SHIFTS):
            key = d + "_" + s
            graph_matrix[1:(1 + df.shape[0]), (1 + df.shape[0] + (i * len(SHIFTS)) + j)] = df[key].to_numpy()
    return graph_matrix


def make_sl_adj_mat(df, graph_matrix_sl, shift_type_dict):
    graph_matrix_sl = graph_matrix_sl.copy()

    for s in shift_type_dict:
        df_sl = None
        for st in shift_type_dict[s]:
            if df_sl is None:
                df_sl = df["sl_" + st].to_numpy().copy()
            else:
                df_sl |= df["sl_" + st].to_numpy()
        if df_sl is None:
            df_sl = np.array(df.shape[0] * [False])
        keysplit = s.split('_')
        column_index = DAYS.index(keysplit[0]) * 3 + SHIFTS.index(keysplit[1])
        graph_matrix_sl[1:(1 + df.shape[0]), 1 + df.shape[0] + column_index][df_sl.transpose() == False] = 0

    graph_matrix_sl[(1 + df.shape[0]):-1, -1] = 1

    return graph_matrix_sl


def make_worker_adj_mat(df, graph_matrix, flow_sl, max_people_at_shift=3):
    graph_matrix = graph_matrix.copy()
    flow_shift_counts = flow_sl.flow.toarray()[(1 + df.shape[0]):-1, -1]
    flow_shift_counts[flow_shift_counts > 0] = max_people_at_shift
    graph_matrix[(1 + df.shape[0]):-1, -1] = flow_shift_counts
    sl_flow_only_pos = flow_sl.flow.toarray().copy()
    sl_flow_only_pos[sl_flow_only_pos < 0] = 0
    graph_matrix = graph_matrix - sl_flow_only_pos

    return graph_matrix


def get_flow_from_adj_mat(adj_mat):
    graph_mat = csr_matrix(adj_mat)
    flow = maximum_flow(graph_mat, 0, adj_mat.shape[0] - 1)
    return flow


def get_shift_list(df, flow_sl, flow, add_flow=None):
    np_flow = flow.flow.toarray()
    np_sl_flow = flow_sl.flow.toarray()
    shift_list = {}
    for i, day in enumerate(DAYS):
        for j, shift in enumerate(SHIFTS):
            key = day + "_" + shift
            ind_people = np.where(np_flow[1:(1 + df.shape[0]), (1 + df.shape[0] + (i * len(SHIFTS)) + j)] == 1)
            ind_sl = np.where(np_sl_flow[1:(1 + df.shape[0]), (1 + df.shape[0] + (i * len(SHIFTS)) + j)] == 1)
            shift_list[key] = {"lead": None, "worker": []}
            if len(ind_sl[0]):
                shift_list[key]["lead"] = df["name"][ind_sl[0][0]]
            for k in list(ind_people[0]):
                shift_list[key]["worker"].append(df['name'][k])
            if add_flow is not None:
                ind_people_2 = np.where(
                    add_flow.flow.toarray()[1:(1 + df.shape[0]), (1 + df.shape[0] + (i * len(SHIFTS)) + j)] == 1)
                for k in list(ind_people_2[0]):
                    shift_list[key]["worker"].append(df['name'][k])
    return shift_list


def get_beer_list(df):
    obj = {}
    beers = df[df["beer"] > 0]
    obj = {}
    for n in list(beers['name']):
        obj[n] = df[df["name"] == n]["beer"]
    return obj


def draw_matching(df, adj_mat, out_file=None):
    G = nx.from_numpy_array(adj_mat)

    node_labels = ["start"]
    node_labels += list(df['name'])
    node_labels += [d + '_' + s for d in DAYS for s in SHIFTS]
    node_labels += ["end"]

    node_labels_obj = {}

    for i, v in enumerate(node_labels):
        node_labels_obj[i] = v

    G = nx.relabel_nodes(G, node_labels_obj)

    pos = []

    pos.append((0, 40))

    for i in range(1, 1 + df.shape[0]):
        pos.append((2, i - 1))

    for i in range(1 + df.shape[0], 1 + df.shape[0] + (len(SHIFTS) * len(DAYS))):
        pos.append((4, (i - 1 - df.shape[0]) * 4))

    pos.append((6, 40))

    pos_obj = {}

    for i, v in enumerate(pos):
        pos_obj[node_labels[i]] = v

    plt.figure(figsize=(15, 40))
    # labels = nx.get_edge_attributes(G,'weight')
    # nx.draw_networkx_edge_labels(G,pos,edge_labels=labels)
    nx.draw_networkx(G, arrows=True, pos=pos_obj)
    if out_file is None:
        plt.show()
    else:
        plt.savefig(out_file)


def main():
    import PySimpleGUI as sg
    from datetime import date
    import os
    import sys

    day_texts = {"mon": "Monday", "tue": "Tuesday", "wed": "Wednesday", "thr": "Thursday", "fri": "Friday",
                 "sat": "Saturday", "sun": "Sunday"}
    time_texts = {"9": "9:00 - 13:00", "14": "14:00 - 18:00", "18": "18:00 - 22:00"}

    layout = [[sg.Column([[sg.Text("Excel file:")], [sg.Text("Sheet name:")]]),
               sg.Column([[sg.InputText(), sg.FileBrowse()], [sg.InputText("ShiftList_KW16")]])],
              [sg.Text("Maximum shift size:"), sg.InputText("3", key="-MAX-SHIFT-SIZE-", size=3)],
              [sg.Button('Generate Shiftplan'), sg.Button('Generate Beerlist'), sg.Text("Comment:"),
               sg.InputText(key="-COMMENT-", size=25)],
              [sg.Multiline(size=(70, 15), key="-OUTBOX-")]]

    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    icon_fqp = os.path.join(base_path, "tufast_img.ico")

    window = sg.Window('TUfast Matching Tool', layout, icon=icon_fqp)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            break

        if event == "Generate Beerlist":
            fname = values[0]
            sname = values[1]
            comment = values["-COMMENT-"]
            df, shift_type_dict = read_sheet(fname, sname)
            beerlist = get_beer_list(df)
            bstr = ""
            for b in beerlist:
                bstr += "|| " + b + """ || Did not fill out shift plan || {X} || """ + date.today().strftime(
                    "%d.%m.%Y") + ' || ' + comment + ' ||\n'
            window["-OUTBOX-"].update(bstr)

        if event == 'Generate Shiftplan':
            fname = values[0]
            sname = values[1]
            if len(sname) == 0 or not os.path.isfile(fname):
                sg.popup("fill valid shiftplan file")
                continue
            if len(sname) == 0:
                sg.popup("fill in sheet name")
                continue
            if len(values["-MAX-SHIFT-SIZE-"]) < 0 or not values["-MAX-SHIFT-SIZE-"].isnumeric():
                sg.popup("fill in maximum shift size")
                continue
            max_shift_size = int(values["-MAX-SHIFT-SIZE-"])
            print("Generate Matching for File:", fname, " Worksheet: ", sname)
            df, shift_type_dict = read_sheet(fname, sname)
            base_mat = shiftplan_to_adj_mat(df)

            shift_degree = [df[s].sum() for s in shift_type_dict]
            len([x for x in shift_degree if x < 2])
            for i,key in enumerate(shift_type_dict):
                if shift_degree[i] < 2:
                    shift_type_dict[key] = []

            # Find the best flow, that results in the most shifts possible.

            best_sl_flow = None
            best_w_mat = None
            best_flow = None
            best_flow_sum_value = 0
            best_df = None

            for _ in range(50):
                df = df.sample(frac=1).reset_index(drop=True)
                base_mat = shiftplan_to_adj_mat(df)
                shift_type_dict_copy = shift_type_dict.copy()
                for _ in range(10):
                    sl_mat = make_sl_adj_mat(df, base_mat, shift_type_dict_copy)
                    flow_sl = get_flow_from_adj_mat(sl_mat)
                    w_mat = make_worker_adj_mat(df, base_mat, flow_sl, 2)
                    flow = get_flow_from_adj_mat(w_mat)

                    shift_list = get_shift_list(df, flow_sl, flow)

                    print(flow_sl.flow_value + flow.flow_value)

                    for k in shift_list:
                        if (shift_list[k]["lead"] is not None) and (
                                shift_list[k] is None or len(shift_list[k]["worker"]) == 0):
                            shift_type_dict_copy[k] = []
                            shift_list = None
                            break

                    if shift_list is not None:
                        break

                if (flow_sl.flow_value + flow.flow_value) > best_flow_sum_value:
                    best_sl_flow = flow_sl
                    best_w_mat = w_mat
                    best_flow = flow
                    best_flow_sum_value = (flow_sl.flow_value + flow.flow_value)
                    best_df = df.copy()

            w_mat_2 = make_worker_adj_mat(best_df, best_w_mat, best_flow, max_shift_size - 1)
            flow_2 = get_flow_from_adj_mat(w_mat_2)
            shift_list = get_shift_list(best_df, best_sl_flow, best_flow, flow_2)
            print(best_sl_flow.flow_value + best_flow.flow_value + flow_2.flow_value)

            sl_str = ""
            for k in shift_list:
                day, time = k.split("_")
                sl_str += day_texts[day] + " " + time_texts[time] + "\n"
                if shift_list[k]["lead"] is None:
                    sl_str += "None\n"
                else:
                    sl_str += "@" + shift_list[k]["lead"] + " (lead)\n"
                    for m in shift_list[k]["worker"]:
                        sl_str += "@" + m + "\n"
                sl_str += "\n"
            window["-OUTBOX-"].update(sl_str)

    window.close()


if __name__ == "__main__":
    main()
