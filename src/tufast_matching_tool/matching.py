import sys
import os
import pandas as pd
import openpyxl
from PySide6.QtWidgets import (QApplication, QMainWindow, QFileDialog, QVBoxLayout, QWidget, QPushButton, QLineEdit, QLabel, QCheckBox, QTextEdit, QFormLayout)
from PySide6.QtCore import Qt
from tufast_matching_tool.general import DAYS, SHIFTS, SHIFT_TYPES
from tufast_matching_tool.maxflow import max_flow_matching
from tufast_matching_tool.mincost import min_cost_matching

START_COL_NAMES = 3
START_ROW_AVAIL = 13
START_ROW_SHIFT_TYPES = 5
START_ROW_MECA = 23
LAST_ROW_MECA = 59




import random

def read_sheet(file="Shiftsplan_lux025.xlsx", sheet="ShiftList_KW16"):
    dataframe = openpyxl.load_workbook(file)
    dataframe1 = dataframe[sheet]
    name = None
    n_shifts = None

    data = {'name': [], 'n_shifts': [], 'beer': [], 'team': []}

    for d in DAYS:
        for s in SHIFTS:
            key = d + "_" + s
            data[key] = []

    for st in SHIFT_TYPES:
        data["sl_" + st] = []

    START_COL_AVAIL_REL = 2
    START_COL_AVAIL = START_COL_NAMES + START_COL_AVAIL_REL
    START_COL_SHIFTLEADS_REL = START_COL_AVAIL_REL + (len(SHIFTS) * len(DAYS))
    START_COL_SHIFTLEADS = START_COL_NAMES + START_COL_SHIFTLEADS_REL
    LAST_COL = START_COL_SHIFTLEADS + len(SHIFT_TYPES) - 1
    LAST_ROW_SHIFT_TYPES = START_ROW_SHIFT_TYPES + len(SHIFT_TYPES) - 1

    shift_type_dict = {}

    end_of_file = False

    for row in range(START_ROW_SHIFT_TYPES, dataframe1.max_row):
        if end_of_file:
            break
        for i, col in enumerate(dataframe1.iter_cols(START_COL_NAMES, LAST_COL)):
            if i == 0 and row >= START_ROW_AVAIL:
                name = col[row].value
                if name is not None and len(name) > 0:
                    data['name'].append(name)
                    if START_ROW_MECA <= row <= LAST_ROW_MECA:
                        data['team'].append('meca')
                    else:
                        data['team'].append('other')
                else:
                    end_of_file = True
                    break
            elif i == 1 and row >= START_ROW_AVAIL:
                if name is not None:
                    n_shifts = col[row].value
                    data['beer'].append(1 if n_shifts is None else 0)
                    if n_shifts is None:
                        n_shifts = 0
                    data['n_shifts'].append(int(n_shifts))
            elif name is not None and i < START_COL_SHIFTLEADS_REL and row >= START_ROW_AVAIL:
                key = DAYS[(i - START_COL_AVAIL_REL) // len(SHIFTS)] + "_" + SHIFTS[(i - START_COL_AVAIL_REL) % len(SHIFTS)]
                if col[row].value is not None:
                    data[key].append(1)
                else:
                    data[key].append(0)
            elif name is not None and row >= START_ROW_AVAIL:
                key = "sl_" + SHIFT_TYPES[i - START_COL_SHIFTLEADS_REL]
                if col[row].value is not None and len(str(col[row].value)) > 0:
                    data[key].append(True)
                else:
                    data[key].append(False)
            elif row < START_ROW_AVAIL and row < LAST_ROW_SHIFT_TYPES and i >= START_COL_AVAIL_REL and i < START_COL_SHIFTLEADS_REL:
                key = DAYS[(i - START_COL_AVAIL_REL) // len(SHIFTS)] + "_" + SHIFTS[(i - START_COL_AVAIL_REL) % len(SHIFTS)]
                shift_type_id = SHIFT_TYPES[row - START_ROW_SHIFT_TYPES]
                if key not in shift_type_dict:
                    shift_type_dict[key] = []
                if col[row].value is not None:
                    shift_type_dict[key].append(shift_type_id)

    df = pd.DataFrame(data)

    meca_df = df[df['team'] == 'meca'].sample(frac=1).reset_index(drop=True)
    other_df = df[df['team'] != 'meca'].sample(frac=1).reset_index(drop=True)
    df_shuffled = pd.concat([meca_df, other_df]).reset_index(drop=True)

    return df_shuffled, shift_type_dict


def get_beer_list(df):
    obj = {}
    beers = df[df["beer"] > 0]
    obj = {}
    for n in list(beers['name']):
        obj[n] = df[df["name"] == n]["beer"]
    return obj

def extract_names_from_shift_list(shift_list):
    names = set()
    for key in shift_list:
        if shift_list[key]['lead']:
            names.add(shift_list[key]['lead'])
        for worker in shift_list[key]['worker']:
            names.add(worker)
    return list(names)
 
def get_not_assigned_names(df, assigned_names):
    not_assigned = df[(df['n_shifts'] >= 1) & (~df['name'].isin(assigned_names))]
    assigned_meca = df[(df['team'] == 'meca') & (df['name'].isin(assigned_names))]

    not_assigned = pd.concat([not_assigned, assigned_meca])
    
    meca_not_assigned = not_assigned[not_assigned['team'] == 'meca']
    other_not_assigned = not_assigned[not_assigned['team'] != 'meca']
    
    return meca_not_assigned['name'].tolist() + other_not_assigned['name'].tolist()




def switch_names(df, assigned_names, not_assigned_names):
    shift_leader_cols = [col for col in df.columns if col.startswith('sl_')]
    shift_leaders = set()
    
    for col in shift_leader_cols:
        leaders = df[df[col] == True]['name']
        shift_leaders.update(leaders)
    
    filtered_assigned_names = [name for name in assigned_names if name not in shift_leaders]
    filtered_not_assigned_names = [name for name in not_assigned_names if name not in shift_leaders]

    
    num_pairs_to_swap = min(len(filtered_assigned_names), len(filtered_not_assigned_names))

    name_to_index = {name: idx for idx, name in enumerate(df['name'])}

    assigned_indices = [name_to_index[name] for name in filtered_assigned_names[:num_pairs_to_swap]]
    not_assigned_indices = [name_to_index[name] for name in filtered_not_assigned_names[:num_pairs_to_swap]]

    assigned_rows = df.loc[assigned_indices].copy()
    not_assigned_rows = df.loc[not_assigned_indices].copy()

    df.loc[assigned_indices, :] = not_assigned_rows.values
    df.loc[not_assigned_indices, :] = assigned_rows.values

    return df


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("TUfast Matching Tool")
        self.setGeometry(100, 100, 800, 600)
        self.initUI()

    def initUI(self):
        layout = QFormLayout()
        
        self.current_file_input = QLineEdit()
        self.current_sheet_input = QLineEdit("ShiftList_KW16")
        self.previous_file_input = QLineEdit()
        self.previous_sheet_input = QLineEdit("ShiftList_KW15")
        self.max_shift_size_input = QLineEdit("3")
        self.use_new_algo_checkbox = QCheckBox("Higher accuracy but slower new algorithm")
        self.comment_input = QLineEdit()
        self.output_text = QTextEdit()

        self.current_file_button = QPushButton("Browse")
        self.previous_file_button = QPushButton("Browse")
        self.generate_shiftplan_button = QPushButton("Generate Shiftplan")
        self.generate_beerlist_button = QPushButton("Generate Beerlist")

        layout.addRow(QLabel("Excel file:"), self.current_file_input)
        layout.addRow(QLabel("Sheet name:"), self.current_sheet_input)
        layout.addRow(self.current_file_button)

        layout.addRow(QLabel("Previous Excel file:"), self.previous_file_input)
        layout.addRow(QLabel("Sheet name:"), self.previous_sheet_input)
        layout.addRow(self.previous_file_button)

        layout.addRow(QLabel("Maximum shift size:"), self.max_shift_size_input)
        layout.addRow(self.use_new_algo_checkbox)
        layout.addRow(QLabel("Comment:"), self.comment_input)

        layout.addRow(self.generate_shiftplan_button, self.generate_beerlist_button)
        layout.addRow(QLabel("Output:"), self.output_text)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        self.current_file_button.clicked.connect(self.browse_current_file)
        self.previous_file_button.clicked.connect(self.browse_previous_file)
        self.generate_shiftplan_button.clicked.connect(self.generate_shiftplan)
        self.generate_beerlist_button.clicked.connect(self.generate_beerlist)

    def browse_current_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.current_file_input.setText(file_name)

    def browse_previous_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.previous_file_input.setText(file_name)

    def generate_beerlist(self):
        try:
            fname = self.current_file_input.text()
            sname = self.current_sheet_input.text()
            df, _ = read_sheet(fname, sname)
            beerlist = get_beer_list(df)
            bstr = "\n".join(beerlist.keys())  
            self.output_text.setPlainText(bstr)
        except Exception as e:
            self.output_text.setPlainText(f"An error occurred: {e}")

        
    def generate_shiftplan(self):
        # try:
            previous_fname = self.previous_file_input.text().strip()
            previous_sname = self.previous_sheet_input.text().strip()
            current_fname = self.current_file_input.text().strip()
            current_sname = self.current_sheet_input.text().strip()

            if not current_sname or not os.path.isfile(current_fname):
                self.output_text.setPlainText("fill valid shiftplan file.")
                return

            if not previous_sname or not os.path.isfile(previous_fname):
                self.output_text.setPlainText("fill valid shiftplan file.")
                return

            if not current_sname:
                self.output_text.setPlainText("fill in sheet name")
                return

            if not previous_sname:
                self.output_text.setPlainText("fill in sheet name")
                return

            max_shift_size_str = self.max_shift_size_input.text().strip()
            if not max_shift_size_str or not max_shift_size_str.isnumeric():
                self.output_text.setPlainText("fill in maximum shift size")
                return

            max_shift_size = int(max_shift_size_str)

            prev_df, prev_shift_type_dict = read_sheet(previous_fname, previous_sname)
            prev_shift_list = max_flow_matching(prev_df, prev_shift_type_dict, max_shift_size)
            assigned_names_prev = extract_names_from_shift_list(prev_shift_list)
            not_assigned_names_prev = get_not_assigned_names(prev_df, assigned_names_prev)

            curr_df, curr_shift_type_dict = read_sheet(current_fname, current_sname)
            curr_df = switch_names(curr_df, assigned_names_prev, not_assigned_names_prev)

            if self.use_new_algo_checkbox.isChecked():
                if max_shift_size != 3:
                    self.output_text.setPlainText("the new algorithm is only implemented for max shiftsize of 3. use the old algo or set shiftsize to 3.")
                    shift_list = []
                else:
                    shift_list = min_cost_matching(curr_df, curr_shift_type_dict)
            else:
                shift_list = max_flow_matching(curr_df, curr_shift_type_dict, max_shift_size)

            if shift_list is None:
                self.output_text.setPlainText("Algorithm didnt find a shift plan ¯\\_(ツ)_/¯")
                return

            sl_str = ""
            shift_count = 0
            worker_count = 0

            day_texts = {"mon": "Monday", "tue": "Tuesday", "wed": "Wednesday", "thr": "Thursday", "fri": "Friday",
                        "sat": "Saturday", "sun": "Sunday"}
            time_texts = {"9": "9:00 - 13:00", "14": "14:00 - 18:00", "18": "18:00 - 22:00"}

            for k in shift_list:
                day, time = k.split("_")
                sl_str += f"{day_texts[day]} {time_texts[time]}\n"

                if shift_list[k]["lead"] is None:
                    sl_str += "None\n"
                else:
                    lead = shift_list[k]["lead"]
                    sl_str += f"@{lead} (lead)\n"
                    shift_count += 1

                    workers = shift_list[k]["worker"]
                    for worker in workers:
                        sl_str += f"@{worker}\n"
                        worker_count += 1

                    sl_str += "\n"

            self.output_text.setPlainText(f"Number of Shifts: {shift_count}\nNumber of Workers: {worker_count}\n{sl_str}")
        # except Exception as e:
        #     self.output_text.setPlainText(f"an Error has occurred: {str(e)}")




def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()